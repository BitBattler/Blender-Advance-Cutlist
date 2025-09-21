bl_info = {
    "name": "Cutter-x-list",
    "author": "Tom StarSky",
    "version": (1, 0),
    "blender": (4, 5, 0),
    "location": "3D-View > N-Panel > Cutlist",
    "description": "Exportiert Cutlist als XLSX/PDF samt Notizen, Orientierung, Verschnittbild und Presets.",
    "category": "Generic",
}

import bpy
from bpy.props import (CollectionProperty, FloatProperty, StringProperty, IntProperty, PointerProperty, BoolProperty, EnumProperty)
from bpy.types import PropertyGroup, UIList, Operator, Panel
from bpy_extras.io_utils import ExportHelper
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import math
import json
import os

SAW_KERF = 4.0  # mm Schnittverlust pro Schnitt

def get_lang(context):
    try:
        return context.scene.cutlist_lang
    except Exception:
        return 'de'

def t(label, context=None):
    tr = {
        "de": {
            "Platten Konfiguration": "Platten Konfiguration",
            "Kommentar": "Kommentar",
            "Ausrichtung": "Ausrichtung",
            "Längs": "Längs",
            "Quer": "Quer",
            "Material für Platte erzeugen": "Material für Platte erzeugen",
            "Cutlist als XLSX exportieren": "Cutlist als XLSX exportieren",
            "Cutlist als PDF exportieren": "Cutlist als PDF exportieren",
            "Preset speichern": "Preset speichern",
            "Preset laden": "Preset laden",
            "Presetname": "Presetname",
            "Schnittbild (Nesting)": "Schnittbild (Nesting)",
            "Schriftgröße": "Schriftgröße",
        },
        "en": {
            "Platten Konfiguration": "Panel Configuration",
            "Kommentar": "Comment",
            "Ausrichtung": "Orientation",
            "Längs": "Longitudinal",
            "Quer": "Crosswise",
            "Material für Platte erzeugen": "Create sheet material",
            "Cutlist als XLSX exportieren": "Export cutlist as XLSX",
            "Cutlist als PDF exportieren": "Export cutlist as PDF",
            "Preset speichern": "Save preset",
            "Preset laden": "Load preset",
            "Presetname": "Preset name",
            "Schnittbild (Nesting)": "Cutting diagram (Nesting)",
            "Schriftgröße": "Font Size",
        }
    }
    lang = get_lang(context) if context else 'de'
    return tr.get(lang, tr['de']).get(label, label)

class PlateItem(PropertyGroup):
    name: StringProperty(name="Plattenname", default="Platte")
    length: FloatProperty(name="Länge (mm)", default=2800.0, min=1.0)
    width: FloatProperty(name="Breite (mm)", default=2070.0, min=1.0)
    thickness: FloatProperty(name="Dicke (mm)", default=18.0, min=1.0)
    comment: StringProperty(name="Kommentar", default="")
    orientation: EnumProperty(
        name="Ausrichtung",
        items=[("LONG", t("Längs"), ""), ("CROSS", t("Quer"), "")],
        default="LONG"
    )

class PlateSettings(PropertyGroup):
    plates: CollectionProperty(type=PlateItem)
    plate_index: IntProperty(default=0)
    preset_name: StringProperty(name="Presetname", default="Standard")
    font_size: IntProperty(name="Schriftgröße", default=18, min=8, max=64)

class CUTLIST_UL_PlateList(UIList):
    bl_idname = "CUTLIST_UL_PlateList"
    def draw_item(self, context, layout, data, item, icon, active_data, active_propname, index):
        if item:
            txt = f"{item.name}: {int(item.length)}x{int(item.width)}x{int(item.thickness)}mm"
            layout.label(text=txt, icon='MESH_GRID')

class CUTLIST_PT_PlatePanel(Panel):
    bl_label = "Platten Konfiguration"
    bl_idname = "CUTLIST_PT_plate_panel"
    bl_space_type = 'VIEW_3D'
    bl_region_type = 'UI'
    bl_category = "Cutlist"

    def draw(self, context):
        layout = self.layout
        platesettings = context.scene.plate_settings
        row = layout.row()
        row.prop(context.scene, "cutlist_lang", expand=True, text=t("Sprache", context))
        row = layout.row()
        row.template_list("CUTLIST_UL_PlateList", "", platesettings, "plates", platesettings, "plate_index")
        col = row.column(align=True)
        col.operator("cutlist.plate_add", icon='ADD', text="")
        col.operator("cutlist.plate_remove", icon='REMOVE', text="")
        if platesettings.plates and platesettings.plate_index < len(platesettings.plates):
            item = platesettings.plates[platesettings.plate_index]
            layout.prop(item, "name", text=t("Plattenname", context))
            layout.prop(item, "length")
            layout.prop(item, "width")
            layout.prop(item, "thickness")
            layout.prop(item, "comment", text=t("Kommentar", context))
            layout.prop(item, "orientation", text=t("Ausrichtung", context))
            layout.operator("cutlist.material_assign", text=t("Material für Platte erzeugen", context))
        layout.separator()
        row = layout.row(align=True)
        row.prop(platesettings, "preset_name", text=t("Presetname", context))
        row = layout.row(align=True)
        row.operator("cutlist.save_preset", icon='CHECKMARK', text=t("Preset speichern", context))
        row.operator("cutlist.load_preset", icon='IMPORT', text=t("Preset laden", context))
        layout.separator()
        layout.prop(platesettings, "font_size", text=t("Schriftgröße", context))
        layout.operator("export_scene.cutlist_xlsx", icon='EXPORT', text=t("Cutlist als XLSX exportieren", context))
        layout.operator("cutlist.export_pdf", icon='DOCUMENTS', text=t("Cutlist als PDF exportieren", context))
        layout.operator("cutlist.nesting_image", icon='UV', text=t("Schnittbild (Nesting)", context))

class CUTLIST_OT_PlateAdd(Operator):
    bl_idname = "cutlist.plate_add"
    bl_label = "Neue Platte hinzufügen"
    def execute(self, context):
        platesettings = context.scene.plate_settings
        new = platesettings.plates.add()
        new.name = f"Platte {len(platesettings.plates)}"
        platesettings.plate_index = len(platesettings.plates)-1
        return {'FINISHED'}

class CUTLIST_OT_PlateRemove(Operator):
    bl_idname = "cutlist.plate_remove"
    bl_label = "Platte entfernen"
    def execute(self, context):
        platesettings = context.scene.plate_settings
        idx = platesettings.plate_index
        if platesettings.plates and idx < len(platesettings.plates):
            platesettings.plates.remove(idx)
            platesettings.plate_index = max(0, idx-1)
        return {'FINISHED'}

class CUTLIST_OT_MaterialAssign(Operator):
    bl_idname = "cutlist.material_assign"
    bl_label = "Material für Platte erzeugen"
    def execute(self, context):
        platesettings = context.scene.plate_settings
        idx = platesettings.plate_index
        if idx < len(platesettings.plates):
            plate = platesettings.plates[idx]
            mat_name = f"{plate.name}_{int(plate.thickness)}mm_{int(plate.length)}x{int(plate.width)}"
            if mat_name not in bpy.data.materials:
                mat = bpy.data.materials.new(mat_name)
                mat.use_nodes = True
                mat.diffuse_color = (0.9, 0.85, 0.7, 1.0)
            self.report({'INFO'}, f"Material '{mat_name}' erstellt!")
        return {'FINISHED'}

class CUTLIST_OT_SavePreset(Operator):
    bl_idname = "cutlist.save_preset"
    bl_label = "Platten-Preset speichern"
    def execute(self, context):
        settings = context.scene.plate_settings
        data = []
        for plate in settings.plates:
            data.append({
                "name": plate.name,
                "length": plate.length,
                "width": plate.width,
                "thickness": plate.thickness,
                "comment": plate.comment,
                "orientation": plate.orientation,
            })
        preset_path = bpy.utils.resource_path('USER') + f"/cutlist_{settings.preset_name}.json"
        with open(preset_path, "w", encoding='utf-8') as fp:
            json.dump(data, fp)
        self.report({'INFO'}, f"Preset gespeichert: {preset_path}")
        return {'FINISHED'}

class CUTLIST_OT_LoadPreset(Operator):
    bl_idname = "cutlist.load_preset"
    bl_label = "Platten-Preset laden"
    def execute(self, context):
        settings = context.scene.plate_settings
        preset_path = bpy.utils.resource_path('USER') + f"/cutlist_{settings.preset_name}.json"
        if os.path.isfile(preset_path):
            with open(preset_path, "r", encoding='utf-8') as fp:
                data = json.load(fp)
            settings.plates.clear()
            for p in data:
                new = settings.plates.add()
                new.name = p["name"]
                new.length = p["length"]
                new.width = p["width"]
                new.thickness = p["thickness"]
                new.comment = p.get("comment", "")
                new.orientation = p.get("orientation", "LONG")
            settings.plate_index = 0
            self.report({'INFO'}, f"Preset geladen: {preset_path}")
        else:
            self.report({'WARNING'}, "Preset existiert nicht!")
        return {'FINISHED'}

class ExportCutlistXLSXOperator(bpy.types.Operator, ExportHelper):
    bl_idname = "export_scene.cutlist_xlsx"
    bl_label = "Export Cutlist als XLSX"
    filename_ext = ".xlsx"

    export_all: BoolProperty(
        name="Export all objects",
        description="Alle Mesh-Objekte exportieren (statt nur selektierte)",
        default=True
    )
    export_sketch: BoolProperty(
        name="Sketch-Objekte exportieren",
        description="Exportiere Objekte, die 'Sketch' im Namen haben",
        default=False
    )

    def draw(self, context):
        layout = self.layout
        layout.prop(self, "export_all")
        layout.prop(self, "export_sketch")

    def execute(self, context):
        settings = context.scene.plate_settings
        wb = Workbook()
        ws = wb.active
        ws.title = "Cutlist"
        ws.append(['PLATTENBEDARF'])
        ws.append(['Dicke (mm)', 'Plattenname', 'Format (mm)', 'Benötigte Platten'])

        if self.export_all:
            objects = [o for o in bpy.context.scene.objects if o.type == 'MESH']
        else:
            objects = [o for o in bpy.context.selected_objects if o.type == 'MESH']
        if not self.export_sketch:
            objects = [o for o in objects if "sketch" not in o.name.lower()]

        parts_grouped = {}
        for obj in objects:
            dims = sorted([obj.dimensions.x, obj.dimensions.y, obj.dimensions.z], reverse=True)
            laenge = round(dims[0],2)
            breite = round(dims[1],2)
            dicke = round(dims[2],2)
            mat_name = obj.active_material.name if getattr(obj, "active_material", None) else ""
            key = (laenge, breite, dicke, mat_name)
            comment = getattr(obj, "comment", "")
            orientation = getattr(obj, "orientation", "LONG")
            if key not in parts_grouped:
                col_name = obj.users_collection[0].name if getattr(obj, "users_collection", []) else 'None'
                parts_grouped[key] = {
                    "name": obj.name,
                    "laenge": laenge,
                    "breite": breite,
                    "dicke": dicke,
                    "col_name": col_name,
                    "mat_name": mat_name,
                    "stueckzahl": 1,
                    "comment": comment,
                    "orientation": orientation,
                }
            else:
                parts_grouped[key]["stueckzahl"] += 1

        headers = [
            'Beispielname', 'Länge (mm)', 'Breite (mm)', 'Dicke (mm)', 'Collection',
            'Plattenmaterial', 'Stückzahl', 'Kommentar', 'Ausrichtung'
        ]
        cutlist_rows = [
            [p["name"], p["laenge"], p["breite"], p["dicke"], p["col_name"], p["mat_name"], p["stueckzahl"], p["comment"], p["orientation"]]
            for p in parts_grouped.values()
        ]

        plates = [p for p in context.scene.plate_settings.plates]

        for plate in plates:
            platemat = f"{plate.name}_{int(plate.thickness)}mm_{int(plate.length)}x{int(plate.width)}"
            benoetigte_platten = 0
            for row in cutlist_rows:
                if row[5] == platemat:
                    teil_l = row[1]
                    teil_b = row[2]
                    qty = row[6]
                    # Nur Orientierung – nicht drehen!
                    if plate.orientation == "CROSS":
                        n_l = int((plate.length + SAW_KERF) // (teil_b + SAW_KERF))
                        n_b = int((plate.width + SAW_KERF) // (teil_l + SAW_KERF))
                    else:  # LONG
                        n_l = int((plate.length + SAW_KERF) // (teil_l + SAW_KERF))
                        n_b = int((plate.width + SAW_KERF) // (teil_b + SAW_KERF))
                    pro_platte = n_l * n_b
                    if pro_platte < 1:
                        pro_platte = 1
                    benoetigte_platten += math.ceil(qty / pro_platte)
            if benoetigte_platten > 0:
                ws.append([plate.thickness, plate.name, f"{int(plate.length)} x {int(plate.width)}", benoetigte_platten])

        ws.append([])
        ws.append(['CUTLIST'])
        ws.append(headers)
        for row in cutlist_rows:
            ws.append(row)
        for col_idx in range(1, ws.max_column+1):
            max_len = 0
            for cell in ws[get_column_letter(col_idx)]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        wb.save(self.filepath)
        self.report({'INFO'}, "Cutlist als XLSX (nur Orientation, keine Drehung) exportiert.")
        return {'FINISHED'}

class CUTLIST_OT_ExportPDF(Operator, ExportHelper):
    bl_idname = "cutlist.export_pdf"
    bl_label = "Export Cutlist als PDF"
    filename_ext = ".pdf"

    def execute(self, context):
        try:
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet

            cutlist_rows = []
            for obj in bpy.context.scene.objects:
                if obj.type == 'MESH':
                    dims = sorted([obj.dimensions.x, obj.dimensions.y, obj.dimensions.z], reverse=True)
                    mat_name = obj.active_material.name if getattr(obj, "active_material", None) else ""
                    cutlist_rows.append([
                        obj.name,
                        int(dims[0]), int(dims[1]), int(dims[2]),
                        obj.users_collection[0].name if obj.users_collection else "",
                        mat_name, 1,
                        getattr(obj, 'comment', ""), getattr(obj, 'orientation', "LONG")
                    ])

            headers = ['Beispielname', 'Länge (mm)', 'Breite (mm)', 'Dicke (mm)', 'Collection',
                       'Plattenmaterial', 'Stückzahl', 'Kommentar', 'Ausrichtung']

            doc = SimpleDocTemplate(self.filepath, pagesize=landscape(A4))
            elements = []
            styles = getSampleStyleSheet()
            elements.append(Paragraph("Cutlist Tabelle (Orientation)", styles["Heading1"]))
            data = [headers] + cutlist_rows
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            elements.append(table)
            doc.build(elements)
            self.report({'INFO'}, "PDF mit Tabelle (nur orientation).")
        except Exception as e:
            self.report({'WARNING'}, f"PDF-Export benötigt reportlab: {e}")
        return {'FINISHED'}

class CUTLIST_OT_NestingImage(Operator, ExportHelper):
    bl_idname = "cutlist.nesting_image"
    bl_label = "Schnittbild exportieren (PNG)"
    filename_ext = ".png"

    def execute(self, context):
        try:
            from PIL import Image, ImageDraw, ImageFont
            settings = context.scene.plate_settings
            font_size = settings.font_size
            if len(settings.plates) == 0:
                self.report({'WARNING'}, "Keine Platte definiert!")
                return {'CANCELLED'}
            plate = settings.plates[settings.plate_index]
            parts = []
            for obj in bpy.context.scene.objects:
                if obj.type == 'MESH':
                    dims = sorted([obj.dimensions.x, obj.dimensions.y], reverse=True)
                    parts.append((int(dims[0]), int(dims[1]), 1, obj.name))
            TARGET_W, TARGET_H = 900, 600
            scale = min(max(plate.length / TARGET_W, plate.width / TARGET_H), plate.length / 300, plate.width / 200)
            pw = int(plate.length / scale)
            ph = int(plate.width / scale)
            pw = max(pw, 300)
            ph = max(ph, 200)
            max_parts = len(parts)
            total_height = ph * max_parts + 32 * max_parts
            img = Image.new("RGBA", (pw, total_height), (255, 255, 255, 255))
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", font_size)
            except Exception:
                font = None
            platenummer = 1
            part_idx = 0
            while part_idx < len(parts):
                top = ph * (platenummer-1) + 32 * (platenummer-1)
                draw.rectangle([0, top, pw, top+ph], outline="black", width=2)
                sheet_label = f"Platte {platenummer} ({int(plate.length)}x{int(plate.width)})"
                if font:
                    draw.text((8, top+8), sheet_label, fill="black", font=font)
                else:
                    draw.text((8, top+8), sheet_label, fill="black")
                x, y, rowh = 0, top + 32, 0
                while part_idx < len(parts):
                    l, b, qty, name = parts[part_idx]
                    # Nur die orientation pro Platte
                    if plate.orientation == "CROSS":
                        w, h = int(b / scale), int(l / scale)
                    else:  # LONG
                        w, h = int(l / scale), int(b / scale)
                    was_rotated = plate.orientation == "CROSS"
                    if x + w > pw:
                        x = 0
                        y += rowh + 6
                        rowh = 0
                    if (y + h) > (top + ph):
                        break
                    fillcol = (190, 230, 245, 255) if was_rotated else (220, 245, 255, 255)
                    text = f"{name}\n{l}x{b}" + (" (quer)" if was_rotated else "")
                    draw.rectangle([x, y, x + w, y + h], outline="black", width=2, fill=fillcol)
                    if font:
                        draw.text((x+4, y+2), text, fill="blue", font=font)
                    else:
                        draw.text((x+4, y+2), text, fill="blue")
                    x += w + 6
                    rowh = max(rowh, h)
                    part_idx += 1
                platenummer += 1
            final_height = ph * (platenummer-1) + 32 * (platenummer-1)
            img = img.crop((0, 0, pw, final_height))
            img.save(self.filepath)
            self.report({'INFO'}, f"{platenummer-1} Platten/Schnittbilder untereinander exportiert (Orientation).")
        except Exception as e:
            self.report({'WARNING'}, f"PIL nötig für PNG-Export: {e}")
        return {'FINISHED'}

classes = (
    PlateItem,
    PlateSettings,
    CUTLIST_UL_PlateList,
    CUTLIST_PT_PlatePanel,
    CUTLIST_OT_PlateAdd,
    CUTLIST_OT_PlateRemove,
    CUTLIST_OT_MaterialAssign,
    ExportCutlistXLSXOperator,
    CUTLIST_OT_SavePreset,
    CUTLIST_OT_LoadPreset,
    CUTLIST_OT_ExportPDF,
    CUTLIST_OT_NestingImage,
)

def register():
    for cls in classes:
        bpy.utils.register_class(cls)
    bpy.types.Scene.plate_settings = PointerProperty(type=PlateSettings)
    bpy.types.Scene.cutlist_lang = EnumProperty(
        name="Sprache",
        description="Sprachauswahl/Language",
        items=[("de", "DE", ""), ("en", "EN", "")],
        default="de"
    )

def unregister():
    for cls in reversed(classes):
        bpy.utils.unregister_class(cls)
    del bpy.types.Scene.plate_settings
    del bpy.types.Scene.cutlist_lang

if __name__ == "__main__":
    register()
